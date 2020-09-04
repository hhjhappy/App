from django.shortcuts import render, redirect, reverse, HttpResponse
from django.http import FileResponse
from . import forms
from . import models
import time
from django.contrib.auth.models import User
from openpyxl import load_workbook


# 随机生成编号
def grantOrderNumber():
    struct_time = time.localtime(time.time())
    number = time.strftime("%Y%m%d%H%M%S", struct_time)
    return number


def index(request):
    if request.method == 'GET':
        number = grantOrderNumber()

        otherInfomrmation = forms.otherforms()
        return render(request, 'ApplicationApp/index.html', locals())
    else:
        number = grantOrderNumber()
        otherInfomrmation = forms.otherforms(request.POST)
        if otherInfomrmation.is_valid():
            other = otherInfomrmation.save(commit=False)
            other.Numbering = number
            other.save()
            request.session['fileid'] = number
            return redirect(reverse('excel'))
        return redirect(reverse('error'))


def success(request):
    return render(request, 'ApplicationApp/success.html', locals())


def error(request):
    return render(request, 'ApplicationApp/error.html', locals())


# def backstage(request):
def export(request):
    nametime = request.session.get('fileid')
    other = models.Other.objects.get(Numbering=nametime)
    allname = nametime + '.xlsx'
    wb = load_workbook('D:\\日常项目\\Application\\static\\template.xlsx')
    ws = wb.active

    ws['B3'] = '所在单位: %s' % other.companyName
    ws['D3'] = '所在部门: %s' % other.departmentName
    ws['C4'] = '%s ' % other.name
    if other.phoneNumber:
        ws['E4'] = '%s' % other.phoneNumber
    if other.telephoneNumber:
        ws['C5'] = '%s' % other.telephoneNumber
    if other.mail:
        ws['E5'] = '%s' % other.mail
    ws['D2'] = '申请编号: %s' % nametime
    ws['E2'] = '申请时间: %s' % str(other.applicationTime)
    typelist = [
        '专线IP',
        '电信专线',
        '联通专线',
        '其他(请填写):'
    ]
    devicelist = [
        '无线路由器',
        '交换机',
        '录像机',
        '无',
        '其他(请填写):',
    ]
    if other.Type == '4':
        ws['B6'] = "%s %s" % (typelist[int(other.Type) - 1], other.typeRemark)
    else:
        ws['B6'] = "%s" % typelist[int(other.Type) - 1]

    if other.otherDevices == '5':
        ws['B7'] = "%s %s" % (devicelist[int(other.otherDevices) - 1], other.DevicesRemark)
    else:
        ws['B7'] = "%s" % devicelist[int(other.otherDevices) - 1]

    ws['B8'] = "%s 至 %s" % (str(other.applicationTime), str(other.theTerm))
    ws['B9'] = "请简要说明专线使用需求（包括使用位置）及用途：\n" \
               "%s" % other.Remarks
    wb.save("%s.xlsx" % nametime)

    file = open(allname, 'rb')
    response = FileResponse(file)
    response['Content-Type'] = 'application/octet-stream'
    response['Content-Disposition'] = 'attachment;filename="%s"' % allname
    return response


import uuid, os


def generateUUID(filename):
    id = str(uuid.uuid4())
    extend = os.path.splitext(filename[1])
    return id + extend


def backstage(request):
    if request.method == 'GET':
        other = models.Other.objects.filter(finish=False)
        back = forms.backstageforms()
        return render(request,'ApplicationApp/backstage.html', locals())
    else:
        back = forms.backstageforms(request.POST,request.FILES)
        finish = request.POST.get('checkbox')
        fin = models.Other.objects.get(id = int(finish))
        if back.is_valid():
            back.save(commit=False)
            back.link = models.Other.objects.get(id = int(finish))
            fin.finish = True
            fin.save()
            back.save()

            return redirect(reverse("success"))
        return redirect(reverse("error"))

def numberlist(request):
    if request.method == 'GET':
        finishinfo = models.Other.objects.filter(finish=True)
        imagename = models.backstage.objects.all()
        return render(request,'ApplicationApp/numberlist.html',locals())